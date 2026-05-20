"""
Interview Analysis Pipeline — v7
=================================
Paste a YouTube link → get an Excel file where every question the interviewer
asked is a column and the condensed, translated answer sits underneath.

Requirements
------------
Install once (run this in your terminal):
    pip install groq openpyxl yt-dlp

System dependency — install ffmpeg separately:
    Windows : winget install ffmpeg
    Mac     : brew install ffmpeg
    Linux   : sudo apt install ffmpeg

You also need a free Groq API key:
    https://console.groq.com → API Keys → Create key

Usage
-----
1. Fill in the CONFIG section below (API key + YouTube link)
2. Run:  python interview_pipeline_v7.py
3. Find your files in the ./output folder when it finishes.
"""

import os
import re
import math
import json
import time
import subprocess
from datetime import datetime
from typing import Optional

from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIG  ←  edit this section
# =============================================================================

# Your free Groq API key — get one at console.groq.com → API Keys → Create key
GROQ_API_KEY = "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"

# YouTube URL of the interview
# Any format works: youtube.com/live/..., youtu.be/..., watch?v=..., shorts/...
YOUTUBE_URL = "https://youtube.com/live/Nv5tB3KBr0Y?feature=share"

# Language for the Excel output (questions, answers, column headers)
# 'ru' = Russian | 'uz' = Uzbek | 'en' = English
OUTPUT_LANGUAGE = "ru"

# Language spoken in the video
# Forcing this prevents Whisper from guessing wrong on quiet sections,
# which caused hallucinations like "Q. Q. Q." in early versions.
# 'ru' = Russian | 'uz' = Uzbek | None = auto-detect
AUDIO_LANGUAGE = "ru"

# Where to save output files — folder is created automatically if it doesn't exist
OUTPUT_DIR = "Downloads/output"

# Path to the folder containing ffmpeg.exe and ffprobe.exe.
# Set to None to rely on PATH. Use an explicit path if yt-dlp can't find ffmpeg.
FFMPEG_DIR = r"C:\Users\HP\ffmpeg\ffmpeg-master-latest-win64-gpl\bin"

# Name of the Excel file (saved inside OUTPUT_DIR)
OUTPUT_EXCEL = "interview_analysis.xlsx"


# =============================================================================
# MODELS — no need to change these
# =============================================================================

# whisper-large-v3-turbo:
#   Groq runs this at 216x real-time on their LPU hardware.
#   A 1-hour interview transcribes in ~15 seconds via API.
#   Much better accuracy than running whisper-small locally, especially for Russian.
WHISPER_MODEL = "whisper-large-v3-turbo"

# llama-3.3-70b-versatile:
#   70B parameter model, the best available on Groq's free tier.
#   Better than smaller models at reconstructing unclear questions and writing
#   actual findings instead of filler like "the respondent mentioned that..."
LLM_MODEL = "llama-3.3-70b-versatile"

LANG_NAMES = {"uz": "Uzbek (O'zbek)", "ru": "Russian", "en": "English"}


# =============================================================================
# SETUP
# =============================================================================

os.environ["GROQ_API_KEY"] = GROQ_API_KEY
os.makedirs(OUTPUT_DIR, exist_ok=True)

gc = Groq(api_key=GROQ_API_KEY)

# Resolve ffmpeg/ffprobe binary paths once at startup.
def _ff(name: str) -> str:
    if FFMPEG_DIR:
        return os.path.join(FFMPEG_DIR, name + (".exe" if os.name == "nt" else ""))
    return name

FFMPEG = _ff("ffmpeg")
FFPROBE = _ff("ffprobe")


# =============================================================================
# SECTION A — AUDIO DOWNLOAD
# =============================================================================

def get_yt_audio(url: str, stem: str = "interview_audio"):
    """
    Download audio from a YouTube video and save it as a 16kHz mono FLAC file.

    Returns: (path, title, duration_seconds)

    Why FLAC at 16kHz mono?
        16kHz is Whisper's native sample rate — higher wastes space.
        Mono halves file size vs stereo. At these settings 8 minutes = ~10-12 MB,
        well under Groq's 25 MB per-request limit.
    """
    url = url.strip()

    # Normalise any YouTube URL format to the standard watch?v= form.
    for pat in [
        r"(?:youtube\.com/live/|youtu\.be/|youtube\.com/shorts/)([\w-]{11})",
        r"(?:v=|/v/)([\w-]{11})",
    ]:
        m = re.search(pat, url)
        if m:
            url = f"https://www.youtube.com/watch?v={m.group(1)}"
            break

    # Fetch title and duration before downloading.
    info = subprocess.run(
        ["yt-dlp", "--print", "%(title)s|||%(duration)s", "--no-playlist", url],
        capture_output=True, text=True,
    )
    title, dur = "Interview", 0
    if "|||" in info.stdout:
        parts = info.stdout.strip().split("|||")
        title = parts[0].strip()
        try:
            dur = int(float(parts[1]))
        except (ValueError, IndexError):
            pass

    print(f"Title    : {title}")
    print(f"Duration : {dur // 60}m {dur % 60}s")
    print("Downloading audio...")

    out_template = os.path.join(OUTPUT_DIR, f"{stem}.%(ext)s")

    # Download audio only (no video), convert to 16kHz mono FLAC.
    dl_cmd = [
        "yt-dlp", "--extract-audio", "--audio-format", "flac",
        "--postprocessor-args", "ffmpeg:-ar 16000 -ac 1",
        "--audio-quality", "0", "--no-playlist", "--no-warnings",
        "-o", out_template,
    ]
    if FFMPEG_DIR:
        dl_cmd += ["--ffmpeg-location", FFMPEG_DIR]
    dl_cmd.append(url)
    dl = subprocess.run(dl_cmd, capture_output=True, text=True)

    if dl.returncode != 0:
        raise RuntimeError(f"Download failed:\n{dl.stderr[-600:]}")

    # yt-dlp sometimes adjusts the extension — find whatever was actually saved.
    path = os.path.join(OUTPUT_DIR, f"{stem}.flac")
    if not os.path.exists(path):
        for f in os.listdir(OUTPUT_DIR):
            if f.startswith(stem):
                path = os.path.join(OUTPUT_DIR, f)
                break

    size_mb = os.path.getsize(path) / 1e6
    print(f"Saved: {os.path.basename(path)}  ({size_mb:.1f} MB)")
    return path, title, dur


def split_audio(path: str, chunk_minutes: int = 8):
    """
    Split a FLAC file into chunks small enough for the Groq API (25 MB limit).

    Returns: (chunks, offsets)
        chunks  — list of file paths, one per chunk
        offsets — list of floats: the logical start time (seconds) of each chunk,
                  not counting the overlap. Used to rebuild absolute timestamps.

    Why 8 minutes?
        8 min × ~1.5 MB/min = ~12 MB per chunk. Comfortable under the 25 MB limit.

    Why 8-second overlap?
        Each chunk starts 8 seconds before its logical start.
        This prevents losing a word that falls exactly on a chunk boundary.
        The transcription step skips the first 8 seconds of chunks after the
        first to avoid duplicating that audio.
    """
    probe = subprocess.run(
        [
            FFPROBE, "-v", "error", "-show_entries", "format=duration",
            "-of", "default=noprint_wrappers=1:nokey=1", path,
        ],
        capture_output=True, text=True,
    )
    total_sec = float(probe.stdout.strip())

    chunk_sec = chunk_minutes * 60
    overlap_sec = 8
    n_chunks = math.ceil(total_sec / chunk_sec)

    if n_chunks == 1:
        return [path], [0.0]

    print(f"Splitting into {n_chunks} chunks of ~{chunk_minutes} min...")
    chunks, offsets = [], []

    for i in range(n_chunks):
        start = max(0, i * chunk_sec - overlap_sec)
        out = os.path.join(OUTPUT_DIR, f"chunk_{i:02d}.flac")
        subprocess.run(
            [
                FFMPEG, "-i", path,
                "-ss", str(start),
                "-t", str(chunk_sec + overlap_sec),
                "-ar", "16000", "-ac", "1",
                out, "-y", "-loglevel", "quiet",
            ],
            capture_output=True,
        )
        size_mb = os.path.getsize(out) / 1e6
        end_min = min(start + chunk_sec, total_sec) / 60
        print(f"  Chunk {i + 1}: {start / 60:.1f}m – {end_min:.1f}m  ({size_mb:.1f} MB)")
        chunks.append(out)
        offsets.append(float(i * chunk_sec))

    return chunks, offsets


# =============================================================================
# SECTION B — TRANSCRIPTION
# =============================================================================

def transcribe_chunk(path: str, language: Optional[str] = None, retries: int = 3):
    """
    Send one audio chunk to Groq Whisper and return the response.

    Returns the Groq response object (has .segments), or None on failure.

    Error handling:
        413 = file too large  → stop immediately, no point retrying
        429 = rate limited    → wait and retry with increasing delay
        Other                 → log and retry
    """
    # Context prompt helps Whisper on domain-specific vocabulary.
    context_prompt = (
        "Это качественное интервью о кредитных картах. "
        "Интервьюер задаёт вопросы, респондент отвечает."
        if (language or "ru") == "ru"
        else ""
    )

    for attempt in range(retries):
        try:
            with open(path, "rb") as f:
                resp = gc.audio.transcriptions.create(
                    file=(os.path.basename(path), f),
                    model=WHISPER_MODEL,
                    language=language,
                    response_format="verbose_json",  # needed for per-segment timestamps
                    temperature=0.0,                 # deterministic output
                    prompt=context_prompt,
                )
            return resp

        except Exception as e:
            err = str(e)
            if "413" in err or "too_large" in err.lower() or "request_too_large" in err.lower():
                print(f"  File too large ({os.path.getsize(path) / 1e6:.1f} MB)")
                return None
            if "429" in err or "rate_limit" in err.lower():
                wait = 20 * (attempt + 1)
                print(f"  Rate limit — waiting {wait}s...")
                time.sleep(wait)
            else:
                print(f"  Transcription error: {err[:200]}")
                return None

    return None


def transcribe_all(chunks, offsets, language: Optional[str] = None):
    """
    Transcribe all audio chunks and merge into one sorted segment list.

    Returns: [{'start': float, 'end': float, 'text': str}, ...]
    Timestamps are absolute (seconds from the start of the full interview).

    Segments are dropped if:
        - no_speech_prob > 0.8  (silence, music, noise)
        - text shorter than 3 characters
        - within the first 8 seconds of a chunk after the first (overlap region)

    The Groq SDK may return segments as objects or plain dicts depending on version.
    The 'get' lambda handles both transparently.
    """
    all_segments = []

    for i, (chunk_path, logical_offset) in enumerate(zip(chunks, offsets), 1):
        size_mb = os.path.getsize(chunk_path) / 1e6
        print(f"Chunk {i}/{len(chunks)} ({size_mb:.1f} MB)...", end=" ", flush=True)
        t0 = time.time()

        result = transcribe_chunk(chunk_path, language)
        if result is None:
            print("failed, skipping.")
            continue

        raw_segs = result.segments if hasattr(result, "segments") else []
        if raw_segs and isinstance(raw_segs[0], dict):
            get = lambda s, k, d=None: s.get(k, d)
        else:
            get = lambda s, k, d=None: getattr(s, k, d)

        added = 0
        for s in raw_segs:
            if get(s, "no_speech_prob", 0) > 0.8:
                continue
            text = (get(s, "text") or "").strip()
            if len(text) < 3:
                continue
            seg_start = get(s, "start", 0)
            if i > 1 and seg_start < 8:
                continue
            all_segments.append({
                "start": round(logical_offset + seg_start, 2),
                "end":   round(logical_offset + get(s, "end", 0), 2),
                "text":  text,
            })
            added += 1

        lang_detected = getattr(result, "language", "?")
        print(f"{added} segments | lang: {lang_detected} | {time.time() - t0:.1f}s")

        if i < len(chunks):
            time.sleep(2)

    all_segments.sort(key=lambda s: s["start"])
    return all_segments


def build_transcript(segs):
    """
    Format segments into numbered, timestamped plain text.

    Output: '001. [00:00.0] text here'

    Line numbers and timestamps help the LLM understand sequence
    and report back where each answer starts.
    """
    lines = []
    for i, s in enumerate(segs):
        m, sc = divmod(s["start"], 60)
        lines.append(f"{i + 1:03d}. [{int(m):02d}:{sc:04.1f}] {s['text']}")
    return "\n".join(lines)


# =============================================================================
# SECTION C — Q&A EXTRACTION
# =============================================================================

def build_qa_prompt(target_lang_name: str) -> str:
    """
    Build the system prompt sent to Llama 3.3 70B.

    Key rules and why they exist:
        '15-25 questions' — without this, the model returned 52 entries from one
            interview by treating every sentence as a question.
        'Merge follow-up questions' — prevents "And why?" becoming its own row.
        Bad/good example — concrete examples work better than abstract rules;
            without them answers were filler like "the respondent mentioned monitoring
            is bad" instead of the actual specific finding.
        'YOU MUST translate EVERY field' — without this strong instruction, the model
            sometimes stops translating mid-session after a rate limit retry.
    """
    return f"""You are analysing a Russian-language qualitative research interview about credit cards in Uzbekistan.

Context:
- Deep-dive interview about credit card usage experience.
- The INTERVIEWER is named Zina. She asks questions and guides the conversation.
- The RESPONDENT is the person being interviewed. They answer.
- The transcript may contain mixed Russian/Uzbek, transliterated text, or noise — work around it.
- Cover the ENTIRE interview from start to finish. Do not skip any section.

Rules:
- Extract only real questions Zina asks. A 45-60 min interview has roughly 15-25 questions.
- Merge follow-up questions on the same topic into one entry.
- key_points: 2-3 sentences of actual content. Write what the respondent said, not
  'respondent mentions that...'. Example: instead of 'The respondent says monitoring is bad'
  write 'The card shows no transaction detail -- only unexplained balance drops are visible.'
- YOU MUST translate EVERY question and key_points into {target_lang_name}.
  Do NOT leave any field untranslated. Output language = {target_lang_name} only.

Return ONLY a JSON array, no markdown, no explanation:
[
  {{
    "question_number": 1,
    "question": "the interviewer's question in Russian",
    "question_translated": "in {target_lang_name}",
    "key_points": "condensed answer in the respondent's language",
    "key_points_translated": "in {target_lang_name}",
    "timestamp_start": "MM:SS",
    "answer_language": "ru/uz/mixed"
  }}
]"""


def chunk_transcript(text: str, max_chars: int = 7000):
    """
    Split the transcript into chunks the LLM can process at once.

    Why 7000 chars? ~1200-1500 words. With the system prompt and max_tokens=2500
    for the response, this fits comfortably in the model's context window.

    Why 5-line overlap? A question near the end of one chunk may have its answer
    at the start of the next. Repeating the last 5 lines bridges that gap.
    """
    if len(text) <= max_chars:
        return [text]

    lines = text.split("\n")
    chunks, cur, cur_len = [], [], 0

    for line in lines:
        cur.append(line)
        cur_len += len(line)
        if cur_len >= max_chars:
            chunks.append("\n".join(cur))
            cur = cur[-5:]
            cur_len = sum(len(l) for l in cur)

    if cur:
        chunks.append("\n".join(cur))

    return chunks


def extract_chunk(chunk: str, n: int, total: int, prompt: str, retries: int = 3):
    """
    Send one transcript chunk to Llama and parse the JSON response.

    Returns a list of Q&A dicts, or [] on failure.

    JSON recovery: if the response is truncated mid-string (model ran out of tokens),
    the parser finds the last complete '}' and closes the array there,
    salvaging partial results rather than losing the whole chunk.
    """
    for attempt in range(retries):
        try:
            resp = gc.chat.completions.create(
                model=LLM_MODEL,
                messages=[
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": f"Chunk {n}/{total}:\n\n{chunk}\n\nJSON only."},
                ],
                temperature=0.1,
                max_tokens=2500,  # lower cap = less truncated JSON
            )
            raw = resp.choices[0].message.content.strip()
            # Strip markdown code fences if the model wrapped the JSON in them.
            raw = re.sub(r"^```json\s*|^```\s*|```\s*$", "", raw, flags=re.MULTILINE).strip()
            # Extract the array even if there is extra text around it.
            m = re.search(r"(\[.*\])", raw, re.DOTALL)
            if m:
                raw = m.group(1)
            # Try to parse. If truncated, recover to last complete object.
            try:
                result = json.loads(raw)
                return result if isinstance(result, list) else []
            except json.JSONDecodeError:
                last = raw.rfind("}")
                if last > 0:
                    try:
                        return json.loads(raw[: last + 1] + "]")
                    except json.JSONDecodeError:
                        pass
            return []

        except Exception as e:
            err = str(e)
            if "429" in err or "rate_limit" in err.lower():
                wait = 25 * (attempt + 1)
                print(f"  Rate limit — waiting {wait}s...")
                time.sleep(wait)
            else:
                print(f"  LLM error: {err[:120]}")
                return []

    return []


def needs_translation(text: str, target_lang_code: str) -> bool:
    """
    Check whether a text field is still in the wrong language.

    Uses Cyrillic character ratio as a heuristic:
        Target Russian → field should have lots of Cyrillic (ratio >= 0.15)
        Target English → field should have very little Cyrillic (ratio <= 0.30)
        Target Uzbek   → mostly Latin, very little Cyrillic (ratio <= 0.10)
    """
    if not text or len(text.strip()) < 5:
        return False
    cyrillic = sum(1 for c in text if "\u0400" <= c <= "\u04FF")
    ratio = cyrillic / len(text)
    if target_lang_code == "ru" and ratio < 0.15:
        return True
    if target_lang_code == "en" and ratio > 0.30:
        return True
    if target_lang_code == "uz" and ratio < 0.10:
        return True
    return False


def repair_translations(pairs, target_lang: str, target_lang_code: str):
    """
    Find any Q&A pairs where the translated fields are in the wrong language
    and re-translate them one at a time.

    This catches cases where rate limit retries caused the LLM to stop
    translating mid-session, which produced mixed-language Excel output
    (e.g. Q1-Q12 in Russian, Q13 onward still in English).
    """
    needs_fix = [
        i for i, p in enumerate(pairs)
        if needs_translation(p.get("question_translated", ""), target_lang_code)
        or needs_translation(p.get("key_points_translated", ""), target_lang_code)
    ]

    if not needs_fix:
        return pairs

    print(f"  Repairing {len(needs_fix)} untranslated pair(s)...")
    repair_prompt = (
        f"Translate the following fields into {target_lang}.\n"
        f'Return ONLY a JSON object with keys "question_translated" and "key_points_translated".\n'
        f"No markdown, no explanation."
    )

    for i in needs_fix:
        p = pairs[i]
        payload = json.dumps(
            {"question": p.get("question", ""), "key_points": p.get("key_points", "")},
            ensure_ascii=False,
        )
        for attempt in range(3):
            try:
                resp = gc.chat.completions.create(
                    model=LLM_MODEL,
                    messages=[
                        {"role": "system", "content": repair_prompt},
                        {"role": "user", "content": payload},
                    ],
                    temperature=0.1,
                    max_tokens=500,
                )
                raw = resp.choices[0].message.content.strip()
                raw = re.sub(r"^```json\s*|^```\s*|```\s*$", "", raw, flags=re.MULTILINE).strip()
                result = json.loads(raw)
                if "question_translated" in result:
                    p["question_translated"] = result["question_translated"]
                if "key_points_translated" in result:
                    p["key_points_translated"] = result["key_points_translated"]
                pairs[i] = p
                print(f"    Fixed Q{p['question_number']}")
                break
            except Exception as e:
                err = str(e)
                if "429" in err or "rate_limit" in err.lower():
                    time.sleep(30 * (attempt + 1))
                else:
                    print(f"    Repair failed Q{p['question_number']}: {err[:80]}")
                    break
        time.sleep(3)

    return pairs


def extract_all_qa(transcript: str):
    """
    Run the full Q&A extraction on a transcript string.

    Chunks the text → calls the LLM on each chunk → deduplicates by question
    text → re-numbers → runs translation repair → returns the final list.
    """
    target_lang_name = LANG_NAMES.get(OUTPUT_LANGUAGE, OUTPUT_LANGUAGE)
    prompt = build_qa_prompt(target_lang_name)
    chunks = chunk_transcript(transcript)

    print(f"Sending {len(chunks)} chunk(s) to {LLM_MODEL}...")

    qa_pairs = []
    seen = set()

    for i, chunk in enumerate(chunks, 1):
        print(f"  chunk {i}/{len(chunks)}...", end=" ", flush=True)
        pairs = extract_chunk(chunk, i, len(chunks), prompt)
        added = 0
        for p in pairs:
            key = " ".join(p.get("question", "").lower().split())[:100]
            if key and key not in seen:
                seen.add(key)
                qa_pairs.append(p)
                added += 1
        print(f"+{added} (total {len(qa_pairs)})")
        if i < len(chunks):
            time.sleep(5)

    for i, p in enumerate(qa_pairs, 1):
        p["question_number"] = i

    # Catch any fields that ended up in the wrong language and fix them.
    qa_pairs = repair_translations(qa_pairs, target_lang_name, OUTPUT_LANGUAGE)

    return qa_pairs


# =============================================================================
# SECTION D — EXCEL EXPORT
# =============================================================================

def make_excel(qa_pairs, title: str, lang_code: str, out_path: str) -> None:
    """
    Build a formatted two-sheet Excel workbook from the extracted Q&A pairs.

    Sheet 1 'Summary':
        Row 1 = title banner spanning all columns
        Row 2 = subtitle (language, question count, date)
        Row 3 = column headers: # | Interview | Timestamp | Language | Q1 | Q2 | ...
        Row 4 = condensed translated answers
        Row 5 = original-language answers (for verification)

    Sheet 2 'Full detail':
        One row per Q&A pair — all fields side by side,
        original language and translated version.
    """
    if not qa_pairs:
        print("No data to export.")
        return

    wb = openpyxl.Workbook()

    # Styles — defined once and reused.
    thin     = Side(style="thin", color="C9D9E8")
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    h_fill   = PatternFill("solid", fgColor="1F5C99")
    b_font   = Font(name="Arial", size=10)
    s_font   = Font(name="Arial", size=9, color="666666", italic=True)
    m_fill   = PatternFill("solid", fgColor="D6E4F0")
    a_fill   = PatternFill("solid", fgColor="EBF3FB")
    w_fill   = PatternFill("solid", fgColor="FFFFFF")
    wrap_top = Alignment(wrap_text=True, vertical="top")
    center   = Alignment(wrap_text=True, horizontal="center", vertical="center")

    lang_label = LANG_NAMES.get(lang_code, lang_code)
    n_q        = len(qa_pairs)
    last_col   = get_column_letter(4 + n_q)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1F3864")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f'{lang_label} output  |  {n_q} questions  |  {datetime.now().strftime("%d %b %Y")}'
    ws["A2"].font      = Font(name="Arial", size=9, color="AAAAAA")
    ws["A2"].fill      = PatternFill("solid", fgColor="1F3864")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    for ci, h in enumerate(["#", "Interview", "Timestamp", "Language"], 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = center; c.border = bdr

    for qi, pair in enumerate(qa_pairs):
        ci     = 5 + qi
        q_text = pair.get("question_translated", pair.get("question", ""))
        c      = ws.cell(row=3, column=ci, value=f"Q{pair['question_number']}.  {q_text}")
        c.font = h_font; c.fill = h_fill; c.alignment = wrap_top; c.border = bdr
    ws.row_dimensions[3].height = 55

    meta_vals = [
        1,
        title,
        qa_pairs[0].get("timestamp_start", ""),
        qa_pairs[0].get("answer_language", "").upper(),
    ]
    for ci, val in enumerate(meta_vals, 1):
        c = ws.cell(row=4, column=ci, value=val)
        c.font      = Font(name="Arial", bold=True, size=10)
        c.fill      = m_fill
        c.alignment = center if ci != 2 else wrap_top
        c.border    = bdr

    for qi, pair in enumerate(qa_pairs):
        c = ws.cell(row=4, column=5 + qi,
                    value=pair.get("key_points_translated", pair.get("key_points", "")))
        c.font = b_font; c.fill = w_fill; c.alignment = wrap_top; c.border = bdr
    ws.row_dimensions[4].height = 90

    # Row 5: original-language answers for comparison.
    ws.merge_cells("A5:D5")
    ws["A5"] = "Original language (reference)"
    ws["A5"].font      = s_font
    ws["A5"].fill      = PatternFill("solid", fgColor="F2F2F2")
    ws["A5"].alignment = center
    ws["A5"].border    = bdr

    for qi, pair in enumerate(qa_pairs):
        c = ws.cell(row=5, column=5 + qi, value=pair.get("key_points", ""))
        c.font = s_font; c.fill = PatternFill("solid", fgColor="F2F2F2")
        c.alignment = wrap_top; c.border = bdr
    ws.row_dimensions[5].height = 70

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    for qi in range(n_q):
        ws.column_dimensions[get_column_letter(5 + qi)].width = 38

    ws.freeze_panes = "E4"

    # ── Sheet 2: Full detail ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Full detail")
    headers2 = [
        "Q#", "Timestamp", "Lang",
        "Question (original)", f"Question ({lang_label})",
        "Key points (original)", f"Key points ({lang_label})",
        "Full answer (original)",
    ]
    widths2 = [5, 10, 7, 32, 32, 32, 32, 45]

    for ci, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = center; c.border = bdr
    ws2.row_dimensions[1].height = 36

    for ri, pair in enumerate(qa_pairs, 2):
        fill = a_fill if ri % 2 == 0 else w_fill
        row_vals = [
            pair.get("question_number", ri - 1),
            pair.get("timestamp_start", ""),
            pair.get("answer_language", "").upper(),
            pair.get("question", ""),
            pair.get("question_translated", ""),
            pair.get("key_points", ""),
            pair.get("key_points_translated", ""),
            pair.get("full_answer", ""),
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = b_font; c.fill = fill
            c.alignment = center if ci <= 3 else wrap_top
            c.border = bdr
        ws2.row_dimensions[ri].height = 75

    for ci, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"  Sheet 1 — Summary    : {n_q} questions as columns")
    print(f"  Sheet 2 — Full detail: original + translated side by side")


# =============================================================================
# MAIN
# =============================================================================

def main() -> None:
    lang_name = LANG_NAMES.get(OUTPUT_LANGUAGE, OUTPUT_LANGUAGE)
    print("=" * 60)
    print("Interview Analysis Pipeline — v7")
    print(f"Output language : {lang_name}")
    print(f"Audio language  : {AUDIO_LANGUAGE or 'auto-detect'}")
    print(f"Output folder   : {os.path.abspath(OUTPUT_DIR)}")
    print("=" * 60)

    # Step 1 — Download
    print("\n[1/5] Downloading audio...")
    audio_path, interview_title, _ = get_yt_audio(YOUTUBE_URL)

    # Step 2 — Split
    print("\n[2/5] Splitting into chunks...")
    audio_chunks, chunk_offsets = split_audio(audio_path)
    print(f"Ready. {len(audio_chunks)} chunk(s) to transcribe.")

    # Step 3 — Transcribe
    print(f"\n[3/5] Transcribing with {WHISPER_MODEL}...")
    segments = transcribe_all(audio_chunks, chunk_offsets, AUDIO_LANGUAGE)
    print(f"Done. {len(segments)} segments total.")

    if not segments:
        print("No segments produced. Check AUDIO_LANGUAGE or the video URL.")
        return

    print("\nFirst few lines:")
    for s in segments[:4]:
        m, sc = divmod(s["start"], 60)
        print(f"  [{int(m):02d}:{sc:04.1f}] {s['text']}")
    print("...")

    # Step 4 — Build transcript
    print("\n[4/5] Building transcript...")
    raw_transcript = build_transcript(segments)

    transcript_path = os.path.join(OUTPUT_DIR, "raw_transcript.txt")
    with open(transcript_path, "w", encoding="utf-8") as f:
        f.write(f"{interview_title}\n{'=' * 60}\n{raw_transcript}")

    print(f"Transcript: {len(segments)} segments, ~{len(raw_transcript.split())} words")
    print(f"Saved to  : {transcript_path}")
    print("(Check this file first if the Excel output looks wrong.)")

    # Step 5 — Extract Q&A
    print(f"\n[5/5] Extracting Q&A with {LLM_MODEL}...")
    qa_pairs = extract_all_qa(raw_transcript)

    if not qa_pairs:
        print("\nNothing extracted. Possible reasons:")
        print("  - Audio quality too low — open raw_transcript.txt and check the text")
        print("  - AUDIO_LANGUAGE does not match the video — change it in CONFIG")
        print("  - Rate limits hit all chunks — wait a few minutes and run again")
        return

    print(f"\nExtracted {len(qa_pairs)} Q&A pairs.")
    print("\n" + "=" * 60)
    for p in qa_pairs:
        print(f"\nQ{p['question_number']} [{p.get('timestamp_start', '?')}]")
        print(f"  Question   : {p.get('question_translated', p.get('question', ''))[:120]}")
        print(f"  Key points : {p.get('key_points_translated', p.get('key_points', ''))[:200]}")
        print("-" * 60)

    # Step 6 — Export
    excel_path = os.path.join(OUTPUT_DIR, OUTPUT_EXCEL)
    print(f"\nExporting Excel to {excel_path}...")
    make_excel(qa_pairs, interview_title, OUTPUT_LANGUAGE, excel_path)

    print(f"\nAll done. Files in: {os.path.abspath(OUTPUT_DIR)}")


if __name__ == "__main__":
    main()